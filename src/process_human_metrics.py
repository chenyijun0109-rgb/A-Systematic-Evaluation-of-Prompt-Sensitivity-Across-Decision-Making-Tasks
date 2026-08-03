from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path
from statistics import mean
from typing import Any


HORIZON_PATH = Path("DATASET/BANDIT/allHorizonData_cut.csv")
IGT_DIR = Path("DATASET/IGT/IGTdataSteingroever2014")
BART_PATH = Path("DATASET/BART/Dataset.xlsx")
DEFAULT_OUTPUT_DIR = Path("outputs/processed/human_metrics")
BART_PARTICIPANT_ID_COLUMN = 0
BART_AGE_COLUMN = 8
BART_MINIMUM_AGE = 18


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return float(text)


def parse_int(value: Any) -> int | None:
    parsed = parse_float(value)
    return int(parsed) if parsed is not None else None


def safe_mean(values: list[float]) -> float:
    return mean(values) if values else 0.0


def optional_mean(values: list[float]) -> float | None:
    """Return None when an event-conditioned metric has no eligible events."""
    return mean(values) if values else None


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def process_horizon(path: Path = HORIZON_PATH) -> list[dict[str, Any]]:
    by_subject: dict[str, list[dict[str, Any]]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            by_subject[row["subjectNumber"]].append(row)

    summaries: list[dict[str, Any]] = []
    for subject, games in sorted(by_subject.items(), key=lambda item: int(item[0])):
        free_choices: list[dict[str, Any]] = []
        first_free_choices: list[dict[str, Any]] = []
        rewards: list[float] = []
        all_choices: list[int] = []

        for game in games:
            game_length = int(game["gameLength"])
            choice_values = [parse_int(game.get(f"c{i}")) for i in range(1, game_length + 1)]
            reward_values = [parse_float(game.get(f"r{i}")) for i in range(1, game_length + 1)]
            valid_rewards = [value for value in reward_values if value is not None]
            rewards.extend(valid_rewards)
            all_choices.extend([choice for choice in choice_values if choice is not None])

            histories = {1: [], 2: []}
            for trial_index in range(1, game_length + 1):
                choice = choice_values[trial_index - 1]
                reward = reward_values[trial_index - 1]
                if choice is None:
                    continue

                if trial_index > 4:
                    observed_mean_1 = safe_mean(histories[1])
                    observed_mean_2 = safe_mean(histories[2])
                    n_observed_1 = len(histories[1])
                    n_observed_2 = len(histories[2])
                    record = {
                        "choice": choice,
                        "game_length": game_length,
                        "information_condition": "unequal_information"
                        if n_observed_1 != n_observed_2
                        else "equal_information",
                        "n_observed_1": n_observed_1,
                        "n_observed_2": n_observed_2,
                        "observed_mean_1": observed_mean_1,
                        "observed_mean_2": observed_mean_2,
                    }
                    free_choices.append(record)
                    if trial_index == 5:
                        first_free_choices.append(record)

                if reward is not None:
                    histories[choice].append(reward)

        summaries.append(
            {
                "task": "horizon",
                "participant_id": subject,
                "n_games": len(games),
                "n_trials": len(rewards),
                "average_reward_per_trial": safe_mean(rewards),
                "exploration_rate": horizon_exploration_rate(free_choices),
                "directed_exploration": horizon_directed_exploration(first_free_choices),
                "horizon_effect": horizon_effect(first_free_choices),
                "switching_rate": switching_rate(all_choices),
            }
        )
    return summaries


def horizon_exploration_rate(records: list[dict[str, Any]]) -> float:
    eligible = [
        record
        for record in records
        if record["observed_mean_1"] != record["observed_mean_2"]
    ]
    if not eligible:
        return 0.0
    exploratory = 0
    for record in eligible:
        best_option = 1 if record["observed_mean_1"] > record["observed_mean_2"] else 2
        if record["choice"] != best_option:
            exploratory += 1
    return exploratory / len(eligible)


def horizon_directed_exploration(records: list[dict[str, Any]]) -> float:
    eligible = [
        record
        for record in records
        if record["information_condition"] == "unequal_information"
    ]
    if not eligible:
        return 0.0
    directed = 0
    for record in eligible:
        less_observed = 1 if record["n_observed_1"] < record["n_observed_2"] else 2
        if record["choice"] == less_observed:
            directed += 1
    return directed / len(eligible)


def horizon_effect(records: list[dict[str, Any]]) -> float:
    horizon_1 = [record for record in records if record["game_length"] == 5]
    horizon_6 = [record for record in records if record["game_length"] == 10]
    return horizon_exploration_rate(horizon_6) - horizon_exploration_rate(horizon_1)


def switching_rate(choices: list[Any]) -> float:
    if len(choices) < 2:
        return 0.0
    switches = sum(1 for previous, current in zip(choices, choices[1:]) if previous != current)
    return switches / (len(choices) - 1)


def process_igt(data_dir: Path = IGT_DIR) -> list[dict[str, Any]]:
    choices = read_subject_matrix(data_dir / "choice_100.csv")
    wins = read_subject_matrix(data_dir / "wi_100.csv")
    losses = read_subject_matrix(data_dir / "lo_100.csv")
    summaries: list[dict[str, Any]] = []

    for subject in sorted(choices):
        subject_choices = [int(value) for value in choices[subject]]
        subject_wins = [float(value) for value in wins[subject]]
        subject_losses = [float(value) for value in losses[subject]]
        n_trials = len(subject_choices)
        advantageous = [choice in (3, 4) for choice in subject_choices]
        deck_counts = {deck: subject_choices.count(index) for deck, index in zip("ABCD", (1, 2, 3, 4))}
        post_loss_trials = []
        switched_after_loss = 0
        for index in range(1, n_trials):
            if subject_losses[index - 1] < 0:
                post_loss_trials.append(index)
                if subject_choices[index] != subject_choices[index - 1]:
                    switched_after_loss += 1

        block_curve: dict[int, int] = {}
        for block in range(1, 6):
            start = (block - 1) * 20
            end = start + 20
            block_curve[block] = sum(1 if value else -1 for value in advantageous[start:end])

        summaries.append(
            {
                "task": "igt",
                "participant_id": subject,
                "n_trials": n_trials,
                "net_score": sum(1 if value else -1 for value in advantageous),
                "advantageous_choice_rate": sum(advantageous) / n_trials if n_trials else 0.0,
                "deck_A_rate": deck_counts["A"] / n_trials if n_trials else 0.0,
                "deck_B_rate": deck_counts["B"] / n_trials if n_trials else 0.0,
                "deck_C_rate": deck_counts["C"] / n_trials if n_trials else 0.0,
                "deck_D_rate": deck_counts["D"] / n_trials if n_trials else 0.0,
                "average_net_outcome": safe_mean(
                    [win + loss for win, loss in zip(subject_wins, subject_losses)]
                ),
                "post_loss_switching_rate": (
                    switched_after_loss / len(post_loss_trials) if post_loss_trials else 0.0
                ),
                "block_wise_learning_curve": json.dumps(block_curve, ensure_ascii=False),
            }
        )
    return summaries


def read_subject_matrix(path: Path) -> dict[str, list[str]]:
    rows: dict[str, list[str]] = {}
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        next(reader)
        for row in reader:
            rows[row[0]] = row[1:]
    return rows


def load_bart_participant_rows(
    path: Path = BART_PATH,
) -> dict[int, list[tuple[Any, ...]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        by_participant: dict[int, list[tuple[Any, ...]]] = defaultdict(list)
        for row in worksheet.iter_rows(values_only=True):
            participant_id = int(row[BART_PARTICIPANT_ID_COLUMN])
            by_participant[participant_id].append(row)
    finally:
        workbook.close()
    return by_participant


def bart_participant_age(participant: int, rows: list[tuple[Any, ...]]) -> int:
    ages = {parse_int(row[BART_AGE_COLUMN]) for row in rows}
    if None in ages or len(ages) != 1:
        raise ValueError(
            f"BART participant {participant} has missing or inconsistent age values: "
            f"{sorted(str(age) for age in ages)}"
        )
    return next(iter(ages))


def bart_filter_summary(
    path: Path = BART_PATH,
    *,
    minimum_age: int = BART_MINIMUM_AGE,
) -> dict[str, Any]:
    by_participant = load_bart_participant_rows(path)
    excluded = [
        (participant, bart_participant_age(participant, rows), len(rows))
        for participant, rows in sorted(by_participant.items())
        if bart_participant_age(participant, rows) < minimum_age
    ]
    included = [
        (participant, rows)
        for participant, rows in sorted(by_participant.items())
        if bart_participant_age(participant, rows) >= minimum_age
    ]
    return {
        "minimum_age": minimum_age,
        "age_column_index_zero_based": BART_AGE_COLUMN,
        "source_participants": len(by_participant),
        "included_participants": len(included),
        "excluded_participants": len(excluded),
        "excluded_participant_ids": [participant for participant, _, _ in excluded],
        "excluded_ages": [age for _, age, _ in excluded],
        "source_rows": sum(len(rows) for rows in by_participant.values()),
        "included_rows": sum(len(rows) for _, rows in included),
        "excluded_rows": sum(row_count for _, _, row_count in excluded),
        "exclusion_reason": "age_below_minimum",
    }


def bart_exclusion_rows(
    path: Path = BART_PATH,
    *,
    minimum_age: int = BART_MINIMUM_AGE,
) -> list[dict[str, Any]]:
    by_participant = load_bart_participant_rows(path)
    return [
        {
            "participant_id": participant,
            "age": bart_participant_age(participant, rows),
            "n_balloons": len(rows),
            "exclusion_reason": f"age_below_{minimum_age}",
        }
        for participant, rows in sorted(by_participant.items())
        if bart_participant_age(participant, rows) < minimum_age
    ]


def process_bart(
    path: Path = BART_PATH,
    *,
    minimum_age: int = BART_MINIMUM_AGE,
) -> list[dict[str, Any]]:
    by_participant = load_bart_participant_rows(path)

    summaries: list[dict[str, Any]] = []
    for participant, rows in sorted(by_participant.items()):
        if bart_participant_age(participant, rows) < minimum_age:
            continue
        pump_counts = [int(row[5]) for row in rows]
        exploded = [bool(row[6]) for row in rows]
        earnings = [float(row[13] or 0.0) for row in rows]
        unexploded_pumps = [pump for pump, did_explode in zip(pump_counts, exploded) if not did_explode]
        post_explosion_changes = [
            pump_counts[index] - pump_counts[index - 1]
            for index in range(1, len(pump_counts))
            if exploded[index - 1]
        ]

        summaries.append(
            {
                "task": "bart",
                "participant_id": participant,
                "n_balloons": len(rows),
                "average_pumps": safe_mean([float(value) for value in pump_counts]),
                "adjusted_average_pumps": safe_mean([float(value) for value in unexploded_pumps]),
                "explosion_rate": sum(exploded) / len(exploded) if exploded else 0.0,
                "average_earning_per_balloon": safe_mean(earnings),
                "post_explosion_adjustment": optional_mean(
                    [float(value) for value in post_explosion_changes]
                ),
            }
        )
    return summaries


def main() -> None:
    parser = argparse.ArgumentParser(description="Process human datasets into metric tables.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    output_dir = args.output_dir
    horizon = process_horizon()
    igt = process_igt()
    bart = process_bart()
    bart_filter = bart_filter_summary()
    bart_exclusions = bart_exclusion_rows()

    write_csv(output_dir / "horizon_human_metrics.csv", horizon)
    write_csv(output_dir / "igt_human_metrics.csv", igt)
    write_csv(output_dir / "bart_human_metrics.csv", bart)
    write_csv(output_dir / "bart_exclusions.csv", bart_exclusions)

    summary = {
        "horizon_participants": len(horizon),
        "igt_participants": len(igt),
        "bart_participants": len(bart),
        "bart_filter": bart_filter,
        "outputs": [
            str(output_dir / "horizon_human_metrics.csv"),
            str(output_dir / "igt_human_metrics.csv"),
            str(output_dir / "bart_human_metrics.csv"),
            str(output_dir / "bart_exclusions.csv"),
        ],
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
